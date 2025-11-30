"""
Export utilities for CSV and Excel (FREE)
"""
import csv
from io import StringIO, BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse


def export_appointments_csv(appointments):
    """
    Export appointments to CSV
    """
    output = StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        'Référence',
        'Utilisateur',
        'Email',
        'Service',
        'Bureau',
        'Date',
        'Heure',
        'Statut',
        'Créé le',
    ])
    
    # Data
    for apt in appointments:
        writer.writerow([
            apt.reference_number,
            apt.user.get_full_name(),
            apt.user.email,
            apt.service_type.name,
            apt.office.name,
            apt.appointment_date.strftime('%d/%m/%Y'),
            apt.appointment_time.strftime('%H:%M'),
            apt.get_status_display(),
            apt.created_at.strftime('%d/%m/%Y %H:%M'),
        ])
    
    return output.getvalue()


def export_applications_excel(applications):
    """
    Export applications to Excel with formatting
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Demandes"
    
    # Header style
    header_fill = PatternFill(start_color="009639", end_color="009639", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Headers
    headers = [
        'Référence',
        'Type',
        'Demandeur',
        'Email',
        'Service',
        'Bureau',
        'Statut',
        'Montant (XOF)',
        'Payé',
        'Soumis le',
        'Complété le',
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row, app in enumerate(applications, 2):
        ws.cell(row=row, column=1, value=app.reference_number)
        ws.cell(row=row, column=2, value=app.get_application_type_display())
        ws.cell(row=row, column=3, value=app.applicant.get_full_name())
        ws.cell(row=row, column=4, value=app.applicant.email)
        ws.cell(row=row, column=5, value=app.service_type.name)
        ws.cell(row=row, column=6, value=app.office.name)
        ws.cell(row=row, column=7, value=app.get_status_display())
        ws.cell(row=row, column=8, value=float(app.total_fee))
        ws.cell(row=row, column=9, value='Oui' if app.is_paid else 'Non')
        ws.cell(row=row, column=10, value=app.submitted_at.strftime('%d/%m/%Y') if app.submitted_at else '')
        ws.cell(row=row, column=11, value=app.completed_at.strftime('%d/%m/%Y') if app.completed_at else '')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


def export_payments_excel(payments):
    """
    Export payments to Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Paiements"
    
    # Header style
    header_fill = PatternFill(start_color="009639", end_color="009639", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Headers
    headers = [
        'Transaction ID',
        'Reçu',
        'Utilisateur',
        'Demande',
        'Montant (XOF)',
        'Devise',
        'Méthode',
        'Statut',
        'Date',
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    total_amount = 0
    for row, payment in enumerate(payments, 2):
        ws.cell(row=row, column=1, value=payment.transaction_id)
        ws.cell(row=row, column=2, value=payment.receipt_number or 'N/A')
        ws.cell(row=row, column=3, value=payment.user.get_full_name())
        ws.cell(row=row, column=4, value=payment.application.reference_number)
        ws.cell(row=row, column=5, value=float(payment.amount))
        ws.cell(row=row, column=6, value=payment.currency)
        ws.cell(row=row, column=7, value=payment.get_payment_method_display())
        ws.cell(row=row, column=8, value=payment.get_status_display())
        ws.cell(row=row, column=9, value=payment.created_at.strftime('%d/%m/%Y %H:%M'))
        
        if payment.status == 'COMPLETED':
            total_amount += float(payment.amount)
    
    # Add total row
    total_row = len(payments) + 3
    ws.cell(row=total_row, column=4, value='TOTAL:').font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=total_amount).font = Font(bold=True)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer

